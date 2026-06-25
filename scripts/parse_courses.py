"""Parse Excel course structure files into JSON data for CourseCenterView."""

import json, re, os, sys
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DIR = os.path.join(BASE, '课程结构')
OUT_DIR = os.path.join(BASE, 'frontend', 'src', 'data')

FILES = [
    ('Python初级1——5章.xlsx',   '初级'),
    ('Python初级6——11章.xlsx',  '初级'),
    ('Python初级12章.xlsx',     '初级'),
    ('python中级1——2章.xlsx',   '中级'),
    ('python中级3章.xlsx',      '中级'),
    ('python中级4——7章.xlsx',   '中级'),
    ('Python高级1-5章.xlsx',    '高级'),
    ('Python高级6——8章.xlsx',   '高级'),
]

def classify_row(val):
    """Return ('chapter'|'section'|'lesson'|None, extracted info)."""
    if not val:
        return (None, None)
    v = val.strip().replace(' ', '').replace('　', '')
    m = re.match(r'第(\d+)章(.+)', v)
    if m:
        return ('chapter', {'num': int(m.group(1)), 'title': m.group(2).strip()})
    m = re.match(r'第[一二三四五六七八九十\d]+节[：:]?(.+)', v)
    if m:
        return ('section', {'title': m.group(1).strip('：: ')})
    m = re.match(r'P(\d+)', v)
    if m:
        return ('lesson', {'page': int(m.group(1))})
    return (None, None)

def extract_bvid(text):
    """Extract bvid from B站 iframe HTML."""
    m = re.search(r'bvid=([A-Za-z0-9]+)', text)
    return m.group(1) if m else None

def parse_duration(val):
    """Parse HH:MM:SS or MM:SS to total seconds."""
    if not val:
        return 0
    s = str(val).strip()
    # Strip trailing junk like " |", " | " that appears in some cells
    s = re.sub(r'\s*\|.*$', '', s)
    # Match time pattern
    m = re.match(r'(\d+):(\d+)(?::(\d+))?', s)
    if not m:
        return 0
    parts = [int(g) for g in m.groups() if g is not None]
    if len(parts) == 3:
        return parts[0] * 3600 + parts[1] * 60 + parts[2]
    elif len(parts) == 2:
        return parts[0] * 60 + parts[1]
    return 0

def format_duration(secs):
    """Format seconds to HH:MM:SS or MM:SS."""
    if secs >= 3600:
        h, r = divmod(secs, 3600)
        m, s = divmod(r, 60)
        return f'{h}:{m:02d}:{s:02d}'
    m, s = divmod(secs, 60)
    return f'{m}:{s:02d}'

def parse_file(filepath, level):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    chapters = []
    current_chapter = None
    current_section = None
    last_bvid = None  # propagate bvid across chapters in same file

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        # Gather all non-empty cell values
        vals = [str(v).strip() if v is not None else '' for v in row]
        # Pads columns to at least 4 for uniform access
        while len(vals) < 4:
            vals.append('')

        col_a = vals[0] if len(vals) > 0 else ''
        col_b = vals[1] if len(vals) > 1 else ''
        col_c = vals[2] if len(vals) > 2 else ''
        col_d = vals[3] if len(vals) > 3 else ''

        rtype, info = classify_row(col_a)

        if rtype == 'chapter':
            current_section = None
            chapter_bvid = None
            # Search all columns for bvid
            for c in vals:
                bv = extract_bvid(c)
                if bv:
                    chapter_bvid = bv
                    last_bvid = bv
                    break
            if not chapter_bvid:
                chapter_bvid = last_bvid  # inherit from previous chapter
            current_chapter = {
                'num': info['num'],
                'title': info['title'],
                'bvid': chapter_bvid,
                'sections': [],
            }
            chapters.append(current_chapter)

        elif rtype == 'section':
            current_section = {
                'title': info['title'],
                'lessons': [],
            }
            if current_chapter:
                current_chapter['sections'].append(current_section)

        elif rtype == 'lesson':
            title = ''
            # Title is usually in col_c; check cols
            for c in [col_c, col_b, col_d]:
                c = c.strip()
                if c and not c.startswith('<iframe') and not c.startswith('//player'):
                    title = c
                    break
            # Clean title: strip leading number prefixes like "01-", "1_"
            title = re.sub(r'^\d+[\.\-_\s]+', '', title).strip()
            duration_sec = parse_duration(col_b)
            lesson = {
                'page': info['page'],
                'title': title,
                'duration': format_duration(duration_sec),
                'durationSec': duration_sec,
            }
            if current_section:
                current_section['lessons'].append(lesson)
            elif current_chapter:
                # No section defined, put directly under chapter
                if not current_chapter['sections']:
                    current_chapter['sections'].append({'title': '', 'lessons': []})
                current_chapter['sections'][0]['lessons'].append(lesson)

    # Filter out empty sections
    for ch in chapters:
        ch['sections'] = [s for s in ch['sections'] if s['lessons']]

    return chapters

def build_course_structure(all_results):
    """Merge files by level into one course structure."""
    levels = {'初级': [], '中级': [], '高级': []}

    for fname, level in FILES:
        filepath = os.path.join(SRC_DIR, fname)
        chapters = parse_file(filepath, level)
        levels[level].extend(chapters)

    # Sort chapters by num within each level
    for level in levels:
        levels[level].sort(key=lambda c: c['num'])

    # Build final structure
    courses = []
    for level_name in ['初级', '中级', '高级']:
        chs = levels[level_name]
        total_lessons = sum(
            sum(len(s['lessons']) for s in ch['sections'])
            for ch in chs
        )
        total_duration_sec = sum(
            sum(l['durationSec'] for s in ch['sections'] for l in s['lessons'])
            for ch in chs
        )
        courses.append({
            'id': level_name,
            'name': f'Python{level_name}',
            'description': f'Python{level_name}课程，共{len(chs)}章{total_lessons}课时',
            'level': level_name,
            'chapterCount': len(chs),
            'lessonCount': total_lessons,
            'totalDuration': format_duration(total_duration_sec),
            'totalDurationSec': total_duration_sec,
            'chapters': chs,
        })

    return courses

def main():
    results = {}
    for fname, level in FILES:
        filepath = os.path.join(SRC_DIR, fname)
        if not os.path.exists(filepath):
            print(f'WARNING: {filepath} not found')
            continue
        chapters = parse_file(filepath, level)
        results[fname] = chapters
        lesson_count = sum(
            sum(len(s['lessons']) for s in ch['sections'])
            for ch in chapters
        )
        print(f'{fname}: {len(chapters)} chapters, {lesson_count} lessons')

    courses = build_course_structure(results)

    # Write JS module
    os.makedirs(OUT_DIR, exist_ok=True)
    js_path = os.path.join(OUT_DIR, 'courseData.js')
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write('// Auto-generated by scripts/parse_courses.py\n')
        f.write('export const courseLevels = ')
        json.dump(courses, f, ensure_ascii=False, indent=2)
        f.write(';\n')

    # Write JSON backup
    json_path = os.path.join(OUT_DIR, 'courseData.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(courses, f, ensure_ascii=False, indent=2)

    print(f'\nWritten: {js_path}')
    print(f'Written: {json_path}')

    # Summary
    for c in courses:
        print(f"  {c['name']}: {c['chapterCount']} chapters, {c['lessonCount']} lessons, {c['totalDuration']}")

if __name__ == '__main__':
    main()
