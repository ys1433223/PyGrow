"""Inspect ALL question bank files - both xlsx and fake-xlsx (CSV)."""
import os
import struct
import csv
import openpyxl

base_dir = r"D:\My_Program\codex\PyGrow\Chapter Resources"

dirs = [
    (os.path.join(base_dir, "python_cer_base_resources", "题库"), "初级"),
    (os.path.join(base_dir, "python_cer_midd_resource", "03-题库"), "中级"),
    (os.path.join(base_dir, "python_cer_senior_resource", "03-题库"), "高级"),
]

for d, stage in dirs:
    if not os.path.exists(d):
        print(f"MISSING: {d}")
        continue
    print(f"\n{'='*80}")
    print(f"STAGE: {stage} | DIR: {d}")
    print(f"{'='*80}")
    for f in sorted(os.listdir(d)):
        if f.startswith("~$"):
            continue
        path = os.path.join(d, f)
        with open(path, "rb") as fh:
            magic = fh.read(10)
        is_zip = magic[:2] == b"PK"

        print(f"\n--- {f} ---")

        if is_zip:
            # Real xlsx
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                print(f"  Type: real xlsx, Sheet={ws.title}, rows={ws.max_row}, cols={ws.max_column}")
                for row_idx in range(1, min(ws.max_row + 1, 5)):
                    cells = []
                    for col in range(1, ws.max_column + 1):
                        v = ws.cell(row_idx, col).value
                        if v is not None:
                            cells.append(str(v)[:120])
                        else:
                            cells.append("(None)")
                    print(f"  Row {row_idx}: {cells}")
                wb.close()
            except Exception as e:
                print(f"  ERROR reading xlsx: {e}")
        else:
            # Likely CSV/TSV with wrong extension
            print(f"  Magic bytes: {magic.hex()}")
            for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'utf-16']:
                try:
                    with open(path, 'r', encoding=enc) as fh:
                        lines = fh.readlines()[:5]
                    print(f"  Encoding {enc}: OK, {len(lines)} lines preview")
                    for i, line in enumerate(lines):
                        print(f"    Line {i+1}: {line.rstrip()[:200]}")
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            else:
                print(f"  Could not decode with any encoding")

print("\n\n=== DONE ===")
