"""
Parse all question bank Excel/CSV files from Chapter Resources into structured data.
Handles multiple formats: real xlsx, CSV with wrong .xlsx extension, multi-row options.
Outputs: questionData.js (frontend) + questions.json (backend seed).
"""
import os
import re
import csv
import json
import openpyxl
from pathlib import Path

BASE_DIR = Path(r"D:\My_Program\codex\PyGrow\Chapter Resources")
OUT_JS = Path(r"D:\My_Program\codex\PyGrow\frontend\src\data\questionData.js")
OUT_JSON = Path(r"D:\My_Program\codex\PyGrow\backend\app\data\questions.json")

STAGE_MAP = {
    "python_cer_base_resources": "初级",
    "python_cer_midd_resource": "中级",
    "python_cer_senior_resource": "高级",
}

# Mappings from Chinese to English
TYPE_MAP = {
    "单选题": "single_choice",
    "单选": "single_choice",
    "多选题": "multiple_choice",
    "多选": "multiple_choice",
    "判断题": "judge",
    "判断": "judge",
    "填空题": "fill_blank",
    "填空": "fill_blank",
    "简答题": "short_answer",
    "简答": "short_answer",
    "代码题": "code",
    "代码": "code",
}

DIFFICULTY_MAP = {
    "易": "easy",
    "简单": "easy",
    "中": "medium",
    "中等": "medium",
    "难": "hard",
    "困难": "hard",
}

# Knowledge tag mapping based on chapter number
CHAPTER_TAG_MAP = {
    1: "Python基础",
    2: "编码规范",
    3: "数据类型",
    4: "运算符与表达式",
    5: "函数",
    6: "正则表达式",
    7: "面向对象",
    8: "文件操作",
    9: "网页基础",
    10: "爬虫基础",
    11: "爬虫进阶",
    12: "数据存储",
}

MID_CHAPTER_TAG_MAP = {
    1: "数据库基础",
    2: "非关系型数据库",
    3: "Django框架",
    4: "Selenium自动化",
    5: "爬虫原理",
    6: "分布式爬虫",
    7: "反爬虫",
}

SENIOR_CHAPTER_TAG_MAP = {
    1: "NumPy科学计算",
    2: "Pandas数据处理",
    3: "数据清洗",
    4: "数据可视化",
    5: "数据分析",
    6: "机器学习",
    7: "深度学习",
    8: "推荐算法",
}


def detect_format(filepath):
    """Detect file format: 'xlsx', 'csv_utf8', 'csv_gbk'."""
    with open(filepath, "rb") as f:
        magic = f.read(4)
    if magic[:2] == b"PK":
        return "xlsx"
    # Try utf-8-sig first (BOM), then gbk
    ext = filepath.suffix.lower()
    if ext == ".csv":
        # Try to detect encoding
        try:
            with open(filepath, "r", encoding="utf-8-sig") as f:
                f.read(100)
            return "csv_utf8"
        except UnicodeDecodeError:
            return "csv_gbk"
    # .xlsx files that aren't zip
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            f.read(100)
        return "csv_utf8"
    except UnicodeDecodeError:
        return "csv_gbk"


def parse_options_string(opt_str):
    """Parse option string like 'A. xxx B. yyy C. zzz' into structured dict."""
    if not opt_str or opt_str == "None" or opt_str == "(None)":
        return []
    # Split on option letters
    # Match patterns like A. or A、 or A) at the start
    parts = re.split(r'\n?(?=[A-E][.、．)）]\s*)', str(opt_str).strip())
    options = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Extract label and text
        m = re.match(r'([A-E])[.、．)）]\s*(.*)', part, re.DOTALL)
        if m:
            options.append({"label": m.group(1), "text": m.group(2).strip()})
    return options


def normalize_type(type_str):
    """Normalize question type string to English enum."""
    if not type_str:
        return "single_choice"
    type_str = str(type_str).strip()
    for cn, en in TYPE_MAP.items():
        if cn in type_str:
            return en
    return "single_choice"


def normalize_difficulty(diff_str):
    """Normalize difficulty string to English enum."""
    if not diff_str:
        return "medium"
    diff_str = str(diff_str).strip()
    for cn, en in DIFFICULTY_MAP.items():
        if cn in diff_str:
            return en
    return "medium"


def extract_chapter_num(chapter_section_str):
    """Extract chapter number from string like '7.1.1 函数的定义和调用' or '1.1 认识Python'."""
    if not chapter_section_str:
        return 1
    s = str(chapter_section_str).strip()
    m = re.match(r'(\d+)', s)
    if m:
        return int(m.group(1))
    return 1


def get_knowledge_tag(stage, chapter_num):
    """Get knowledge tag based on stage and chapter number."""
    if stage == "初级":
        return CHAPTER_TAG_MAP.get(chapter_num, f"第{chapter_num}章")
    elif stage == "中级":
        return MID_CHAPTER_TAG_MAP.get(chapter_num, f"第{chapter_num}章")
    else:
        return SENIOR_CHAPTER_TAG_MAP.get(chapter_num, f"第{chapter_num}章")


def parse_csv_file(filepath, stage, encoding):
    """Parse a CSV question bank file."""
    questions = []
    with open(filepath, "r", encoding=encoding) as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return questions

    headers = [h.strip() if h else "" for h in rows[0]]
    # Build column index
    col_map = {}
    for i, h in enumerate(headers):
        h_clean = h.replace(" ", "").replace("﻿", "")
        if "题目" in h_clean and "选项" not in h_clean and "答案" not in h_clean:
            col_map["question"] = i
        elif "题型" in h_clean or "类型" in h_clean:
            col_map["type"] = i
        elif "难度" in h_clean:
            col_map["difficulty"] = i
        elif "章节" in h_clean or "所属" in h_clean:
            col_map["chapter"] = i
        elif "选项" in h_clean and "A" not in h_clean and "B" not in h_clean:
            col_map["options"] = i
        elif "答案" in h_clean or "正确" in h_clean:
            col_map["answer"] = i
        elif "解析" in h_clean or "解释" in h_clean:
            col_map["analysis"] = i
        elif "代码" in h_clean or "预期" in h_clean or "测试" in h_clean:
            col_map["code_ref"] = i

    # Default column mappings for common patterns
    if len(headers) >= 8 and "question" not in col_map:
        # Pattern: 序号,题目,题型,难度,所属章节,选项,答案,解析
        if "序号" in headers[0] or "编号" in headers[0]:
            col_map = {"question": 1, "type": 2, "difficulty": 3, "chapter": 4,
                       "options": 5, "answer": 6, "analysis": 7}
        else:
            col_map = {"question": 0, "type": 1, "difficulty": 2, "chapter": 3,
                       "options": 4, "answer": 5, "analysis": 6}

    chapter_num = 1
    for row in rows[1:]:
        if not row or all(not c or str(c).strip() == "" for c in row):
            continue

        def get_col(key):
            if key in col_map and col_map[key] < len(row):
                v = row[col_map[key]]
                return str(v).strip() if v else ""
            return ""

        question_text = get_col("question")
        if not question_text or question_text == "None":
            continue

        type_str = get_col("type")
        diff_str = get_col("difficulty")
        chapter_str = get_col("chapter")
        opt_str = get_col("options")
        answer_str = get_col("answer")
        analysis_str = get_col("analysis")
        code_ref = get_col("code_ref")

        chapter_num = extract_chapter_num(chapter_str)
        q_type = normalize_type(type_str)
        difficulty = normalize_difficulty(diff_str)
        options = parse_options_string(opt_str)
        knowledge_tag = get_knowledge_tag(stage, chapter_num)

        q = {
            "stage": stage,
            "chapter": chapter_str,
            "chapter_num": chapter_num,
            "knowledge_tag": knowledge_tag,
            "question_type": q_type,
            "difficulty": difficulty,
            "question": question_text,
            "options": options,
            "answer": answer_str,
            "analysis": analysis_str,
            "test_cases": [],
            "starter_code": "",
            "score": 5 if difficulty == "easy" else (10 if difficulty == "medium" else 15),
        }

        # Handle code reference / test cases
        if code_ref and code_ref not in ("None", "", "无"):
            # Try to parse as JSON test_cases first
            try:
                tc = json.loads(code_ref)
                if isinstance(tc, list):
                    q["test_cases"] = tc
                else:
                    q["starter_code"] = code_ref
            except (json.JSONDecodeError, TypeError):
                q["starter_code"] = code_ref

        # For judge questions, normalize answer
        if q_type == "judge":
            ans = answer_str.strip()
            if ans in ("正确", "对", "True", "true", "√", "✓", "是"):
                q["answer"] = "正确"
            elif ans in ("错误", "错", "False", "false", "×", "✗", "否"):
                q["answer"] = "错误"

        questions.append(q)

    return questions


def parse_xlsx_standard(ws, stage):
    """Parse a standard xlsx where questions are in rows with options in one cell or columns."""
    questions = []
    max_row = ws.max_row
    max_col = ws.max_column

    # Find header row (look for 题目 in first 3 rows)
    header_row = 1
    for r in range(1, min(4, max_row + 1)):
        for c in range(1, max_col + 1):
            v = str(ws.cell(r, c).value or "")
            if "题目" in v and "选项" not in v:
                header_row = r
                break
        if header_row != 1:
            break

    # Map columns from header
    headers = []
    for c in range(1, max_col + 1):
        v = str(ws.cell(header_row, c).value or "")
        headers.append(v.replace(" ", ""))

    col_map = {}
    # Check if options are in separate columns (选项A, 选项B pattern)
    has_split_options = any("选项A" in h or "选项A" in h for h in headers)

    for i, h in enumerate(headers):
        if "题目" in h and "选项" not in h and "答案" not in h:
            col_map["question"] = i
        elif "题型" in h or "类型" in h:
            col_map["type"] = i
        elif "难度" in h:
            col_map["difficulty"] = i
        elif "章节" in h or "所属" in h:
            col_map["chapter"] = i
        elif "选项" in h and not has_split_options:
            col_map["options"] = i
        elif "选项A" in h:
            col_map["optA"] = i
        elif "选项B" in h:
            col_map["optB"] = i
        elif "选项C" in h:
            col_map["optC"] = i
        elif "选项D" in h:
            col_map["optD"] = i
        elif "选项E" in h:
            col_map["optE"] = i
        elif "答案" in h or "正确" in h:
            col_map["answer"] = i
        elif "解析" in h or "解释" in h:
            col_map["analysis"] = i
        elif "代码" in h or "预期" in h or "测试" in h:
            col_map["code_ref"] = i

    # Fallback for known patterns
    if not col_map:
        if has_split_options:
            col_map = {"question": 0, "type": 1, "difficulty": 2, "chapter": 3,
                       "optA": 4, "optB": 5, "optC": 6, "optD": 7, "answer": -2, "analysis": -1}
        else:
            col_map = {"question": 0, "type": 1, "difficulty": 2, "chapter": 3,
                       "options": 4, "answer": 5, "analysis": 6}

    chapter_num = 1
    for row_idx in range(header_row + 1, max_row + 1):
        def get_col(key):
            idx = col_map.get(key, -1)
            if 0 <= idx < max_col:
                v = ws.cell(row_idx, idx + 1).value
                return str(v).strip() if v is not None else ""
            return ""

        question_text = get_col("question")
        if not question_text or question_text == "None" or question_text.startswith("Python"):
            continue

        type_str = get_col("type")
        diff_str = get_col("difficulty")
        chapter_str = get_col("chapter")
        answer_str = get_col("answer")
        analysis_str = get_col("analysis")
        code_ref = get_col("code_ref")

        # Build options
        if has_split_options:
            opts = []
            for label in ["A", "B", "C", "D", "E"]:
                key = f"opt{label}"
                if key in col_map:
                    txt = get_col(key)
                    if txt and txt != "None":
                        opts.append({"label": label, "text": txt})
            options = opts
        else:
            opt_str = get_col("options")
            options = parse_options_string(opt_str)

        chapter_num = extract_chapter_num(chapter_str)
        q_type = normalize_type(type_str)
        difficulty = normalize_difficulty(diff_str)
        knowledge_tag = get_knowledge_tag(stage, chapter_num)

        q = {
            "stage": stage,
            "chapter": chapter_str,
            "chapter_num": chapter_num,
            "knowledge_tag": knowledge_tag,
            "question_type": q_type,
            "difficulty": difficulty,
            "question": question_text,
            "options": options,
            "answer": answer_str,
            "analysis": analysis_str,
            "test_cases": [],
            "starter_code": "",
            "score": 5 if difficulty == "easy" else (10 if difficulty == "medium" else 15),
        }

        if code_ref and code_ref not in ("None", "", "无"):
            try:
                tc = json.loads(code_ref)
                if isinstance(tc, list):
                    q["test_cases"] = tc
                else:
                    q["starter_code"] = code_ref
            except (json.JSONDecodeError, TypeError):
                q["starter_code"] = code_ref

        if q_type == "judge":
            ans = answer_str.strip()
            if ans in ("正确", "对", "True", "true", "√", "✓", "是"):
                q["answer"] = "正确"
            elif ans in ("错误", "错", "False", "false", "×", "✗", "否"):
                q["answer"] = "错误"

        questions.append(q)

    return questions


def parse_xlsx_multiline_options(ws, stage):
    """
    Parse the bizarre format where options are split across multiple rows:
    Row N:   序号, 题目(stem), 题型, 难度, 所属章节, 答案, 解析
    Row N+1: (None), A. option text, (None), ... (just the option label+text in col 2)
    Row N+2: (None), B. option text, ...
    etc.

    The question stem is in col 1 (0-indexed), and the options are in col 1 of subsequent rows.
    """
    questions = []
    max_row = ws.max_row
    max_col = ws.max_column

    row = 1
    # Skip header
    while row <= max_row:
        v0 = ws.cell(row, 1).value  # 序号 or question number
        v1 = ws.cell(row, 2).value  # 题目 or option text
        v2 = ws.cell(row, 3).value  # 题型
        v3 = ws.cell(row, 4).value  # 难度
        v4 = ws.cell(row, 5).value  # 所属章节
        v5 = ws.cell(row, 6).value  # 答案
        v6 = ws.cell(row, 7).value  # 解析

        # Check if this row is a header
        if v1 and str(v1).strip() == "题目":
            row += 1
            continue

        # Check if this row is a question stem (non-None 序号 or has 题型/难度)
        v0_str = str(v0).strip() if v0 is not None else ""
        v1_str = str(v1).strip() if v1 is not None else ""

        # A question row has a number in col 0 or has 题型 in col 2
        is_question = False
        if v0_str and v0_str != "None" and v0_str != "序号":
            try:
                int(v0_str)
                is_question = True
            except ValueError:
                pass

        # Also check if col 3 (题型) has a type string
        v2_str = str(v2).strip() if v2 is not None else ""
        if not is_question and v2_str and any(t in v2_str for t in TYPE_MAP):
            is_question = True

        if not is_question:
            row += 1
            continue

        # This is a question stem row
        question_text = v1_str
        if not question_text:
            row += 1
            continue

        type_str = str(v2).strip() if v2 else ""
        diff_str = str(v3).strip() if v3 else ""
        chapter_str = str(v4).strip() if v4 else ""
        answer_str = str(v5).strip() if v5 else ""
        analysis_str = str(v6).strip() if v6 else ""

        # Now scan following rows for options (col 0 = None, col 1 has option text)
        options = []
        next_row = row + 1
        while next_row <= max_row:
            n0 = ws.cell(next_row, 1).value  # Should be None
            n1 = ws.cell(next_row, 2).value  # Option text like "A. xxx" or "B. yyy"
            n2 = ws.cell(next_row, 3).value  # Should be None
            n1_str = str(n1).strip() if n1 is not None else ""

            # Check if next row is a new question (has number in col 0 or 题型 in col 3)
            n0_str = str(n0).strip() if n0 is not None else ""
            n2_str = str(n2).strip() if n2 is not None else ""
            if n0_str and n0_str != "None":
                try:
                    int(n0_str)
                    break  # Next question
                except ValueError:
                    pass
            if n2_str and any(t in n2_str for t in TYPE_MAP):
                break  # Next question

            # This is an option row
            m = re.match(r'([A-E])[.、．)）]\s*(.*)', n1_str)
            if m:
                options.append({"label": m.group(1), "text": m.group(2).strip()})

            next_row += 1

        chapter_num = extract_chapter_num(chapter_str)
        q_type = normalize_type(type_str)
        difficulty = normalize_difficulty(diff_str)
        knowledge_tag = get_knowledge_tag(stage, chapter_num)

        q = {
            "stage": stage,
            "chapter": chapter_str,
            "chapter_num": chapter_num,
            "knowledge_tag": knowledge_tag,
            "question_type": q_type,
            "difficulty": difficulty,
            "question": question_text,
            "options": options,
            "answer": answer_str,
            "analysis": analysis_str,
            "test_cases": [],
            "starter_code": "",
            "score": 5 if difficulty == "easy" else (10 if difficulty == "medium" else 15),
        }

        if q_type == "judge":
            ans = answer_str.strip()
            if ans in ("正确", "对", "True", "true", "√", "✓", "是"):
                q["answer"] = "正确"
            elif ans in ("错误", "错", "False", "false", "×", "✗", "否"):
                q["answer"] = "错误"

        questions.append(q)
        row = next_row  # Skip past the options we just read

    return questions


def parse_file(filepath, stage):
    """Parse a single question bank file, auto-detecting format."""
    fmt = detect_format(filepath)
    ext = filepath.suffix.lower()

    if fmt == "xlsx":
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Determine if this is multi-line options format:
        # Check if row 2 col 0 has a number and row 3 col 0 is None but row 3 col 1 has option text
        is_multiline = False
        if ws.max_row > 2:
            r2_c0 = ws.cell(2, 1).value
            r3_c0 = ws.cell(3, 1).value
            r3_c1 = ws.cell(3, 2).value
            r2_str = str(r2_c0).strip() if r2_c0 is not None else ""
            r3_str = str(r3_c0).strip() if r3_c0 is not None else ""
            r3c1_str = str(r3_c1).strip() if r3_c1 is not None else ""
            if r2_str.isdigit() and (r3_str == "None" or r3_str == "") and r3c1_str.startswith(("A.", "A）", "A)", "A、")):
                is_multiline = True

        if is_multiline:
            questions = parse_xlsx_multiline_options(ws, stage)
        else:
            questions = parse_xlsx_standard(ws, stage)

        wb.close()
        return questions
    else:
        # CSV file
        enc = "utf-8-sig" if fmt == "csv_utf8" else "gbk"
        return parse_csv_file(filepath, stage, enc)


def main():
    all_questions = []
    stats = {}

    # Walk through the 3 directories
    for resource_dir, stage in STAGE_MAP.items():
        stage_path = BASE_DIR / resource_dir
        # Find 题库 directory
        for subdir in stage_path.iterdir():
            if subdir.is_dir() and "题库" in subdir.name:
                qbank_dir = subdir
                break
        else:
            # Try direct 03-题库 pattern
            qbank_dir = stage_path / "03-题库"
            if not qbank_dir.exists():
                print(f"WARNING: No question bank dir found in {stage_path}")
                continue

        file_count = 0
        for f in sorted(qbank_dir.iterdir()):
            if f.name.startswith("~$"):
                continue
            if f.suffix.lower() in (".xlsx", ".xls", ".csv"):
                try:
                    questions = parse_file(f, stage)
                    all_questions.extend(questions)
                    file_count += 1
                    print(f"  {stage} / {f.name}: {len(questions)} questions")
                except Exception as e:
                    print(f"  ERROR {stage} / {f.name}: {e}")
                    import traceback
                    traceback.print_exc()

        stats[stage] = file_count

    # Assign unique IDs
    for i, q in enumerate(all_questions):
        q["question_id"] = i + 1

    # Summary
    print(f"\n{'='*60}")
    print(f"Total questions parsed: {len(all_questions)}")
    for stage, count in stats.items():
        stage_qs = [q for q in all_questions if q["stage"] == stage]
        print(f"  {stage}: {len(stage_qs)} questions from {count} files")
        # By type
        types = {}
        for q in stage_qs:
            t = q["question_type"]
            types[t] = types.get(t, 0) + 1
        print(f"    Types: {types}")
        # By difficulty
        diffs = {}
        for q in stage_qs:
            d = q["difficulty"]
            diffs[d] = diffs.get(d, 0) + 1
        print(f"    Difficulty: {diffs}")

    # Generate output files
    os.makedirs(OUT_JS.parent, exist_ok=True)
    os.makedirs(OUT_JSON.parent, exist_ok=True)

    # Write JS data file for frontend
    js_content = f"""// Auto-generated question bank data
// Total: {len(all_questions)} questions across 3 levels
// Generated by scripts/parse_questions.py

export const questionBank = {json.dumps(all_questions, ensure_ascii=False, indent=2)};

// Helper: get questions by stage
export function getQuestionsByStage(stage) {{
  return questionBank.filter(q => q.stage === stage);
}}

// Helper: get questions by knowledge tag
export function getQuestionsByTag(tag) {{
  return questionBank.filter(q => q.knowledge_tag === tag);
}}

// Helper: get questions by chapter number and stage
export function getQuestionsByChapter(stage, chapterNum) {{
  return questionBank.filter(q => q.stage === stage && q.chapter_num === chapterNum);
}}

// Helper: get all unique knowledge tags
export function getAllKnowledgeTags() {{
  return [...new Set(questionBank.map(q => q.knowledge_tag))];
}}

// Helper: get all unique chapter numbers by stage
export function getChaptersByStage(stage) {{
  return [...new Set(
    questionBank.filter(q => q.stage === stage).map(q => q.chapter_num)
  )].sort((a, b) => a - b);
}}
"""
    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(js_content)
    print(f"\nWrote JS data: {OUT_JS} ({len(js_content)} bytes)")

    # Write JSON for backend
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(all_questions, f, ensure_ascii=False, indent=2)
    print(f"Wrote JSON data: {OUT_JSON}")


if __name__ == "__main__":
    main()
